import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
import win32com.client as win32
import sys
import os
from pathlib import Path
from pivot_table import run_excel


win32c = win32.constants
f_path = Path.cwd()
f_name = 'result.xlsx'


#######################################################
#Добавление столбца reopens на исходный (активный) лист
#######################################################
def insert_iter_col():
    wb = load_workbook('test.xlsx')
    sheet = wb.active
    sheet.insert_cols(2, 1)
    sheet['B1'] = 'iterations new'
    sheet['B2'] = 1
    wb.save('result.xlsx')
    print('>>>added "iterations_new" col')


##################################################################
#Заполнение столбца iterations_new значениями, формула смотрит на chat_id
##################################################################
def fill_iter():
    wb = load_workbook('result.xlsx')
    sheet = wb.active
    max_rows = sheet.max_row
    for i in range(3, max_rows):
        result = np.where(
            sheet[f'A{i}'].value == sheet[f'A{i-1}'].value, sheet[f'B{i-1}'].value + 1, 1) #Формула - =ЕСЛИ(А3=А2;B2+1;1)
        sheet[f'B{i}'] = int(result)
    wb.save('result.xlsx')
    print('>>>"iterations_new" filled')


#####################################################################
#Добавление столбца answered(Есть ответ?) на исходный (активный) лист
#####################################################################
def insert_response_col():
    wb = load_workbook('result.xlsx')
    sheet = wb.active
    sheet.insert_cols(8, 1)
    sheet['H1'] = 'Есть ответ'
    sheet.insert_cols(9, 1)
    sheet['I1'] = 'Нет ответа'
    wb.save('result.xlsx')
    print('>>>added "response" col')


########################################################################################
#Заполнение столбца answered значениями, формула смотрит на столбец response из выгрузки
########################################################################################
def fill_response():
    wb = load_workbook('result.xlsx')
    sheet = wb.active
    max_rows = sheet.max_row
    for i in range(2, max_rows):
        result = np.where(sheet[f'G{i}'].value == None, '0', '1')
        sheet[f'H{i}'] = str(result)
    for i in range(2, max_rows):
        result = np.where(sheet[f'G{i}'].value == None, '1', '0') 
        sheet[f'I{i}'] = str(result)
    wb.save('result.xlsx')
    print('>>>"response" filled')


def moving_cols():
    wb = load_workbook('result.xlsx')
    sheet = wb.active
    max_rows = sheet.max_row
    sheet.insert_cols(10, 1)
    sheet.move_range(f'FB1:FB{max_rows}', rows=0, cols=-148)
    sheet.insert_cols(11, 1)
    sheet.move_range(f'CO1:CO{max_rows}', rows=0, cols=-82)
    sheet.delete_cols(93, 1)
    sheet.delete_cols(158, 1)
    wb.save('result.xlsx')
    print('moving cols done <3')



if __name__ == '__main__':
    insert_iter_col()
    fill_iter()
    insert_response_col()
    fill_response()
    moving_cols()
    run_excel(f_path, f_name)